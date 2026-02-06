import re
import shutil
from pathlib import Path
from datetime import datetime
from typing import Optional, Union, List, Tuple, Dict, Any
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 设置日志
logger = logging.getLogger(__name__)


def is_file_within_workspace(file_path: Path, workspace_path: Path) -> bool:
    """检查文件是否在工作区内"""
    try:
        return str(file_path.resolve()).startswith(str(workspace_path.resolve()))
    except Exception:
        return False


def read_excel_file(file_path: Union[str, Path],
                    workspace_path: Optional[Union[str, Path]] = None,
                    sheet_name: Optional[str] = None) -> Tuple[bool, Dict]:
    """读取Excel文件内容

    Args:
        file_path: Excel文件路径
        workspace_path: 工作区路径
        sheet_name: 工作表名称，None则读取第一个工作表

    Returns:
        Tuple[bool, Dict]: (是否成功, 包含工作簿信息)
    """
    try:
        # 处理文件路径
        if isinstance(file_path, str):
            file_path = Path(file_path)

        # 处理工作区路径
        if workspace_path is not None:
            workspace = Path(workspace_path)
        else:
            workspace = Path.cwd()

        # 处理相对路径
        if not file_path.is_absolute():
            file_path = workspace / file_path

        # 安全检查
        if not is_file_within_workspace(file_path, workspace):
            return False, {"error": "文件不在当前工作区内"}

        if not file_path.exists():
            return False, {"error": f"文件不存在: {file_path}"}

        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            return False, {"error": "文件不是Excel格式"}

        # 加载工作簿
        wb = load_workbook(filename=file_path, data_only=True)

        # 获取工作表
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                return False, {"error": f"工作表 '{sheet_name}' 不存在"}
        else:
            ws = wb.active

        # 读取工作表数据
        data = []
        max_row = ws.max_row
        max_column = ws.max_column

        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            data.append(row_data)

        # 获取工作表信息
        sheets_info = []
        for sheet in wb.sheetnames:
            sheet_ws = wb[sheet]
            sheets_info.append({
                "name": sheet,
                "active": sheet == ws.title,
                "max_row": sheet_ws.max_row,
                "max_column": sheet_ws.max_column
            })

        result = {
            "file_path": str(file_path),
            "active_sheet": ws.title,
            "max_row": max_row,
            "max_column": max_column,
            "data": data,
            "sheets": sheets_info,
            "sheet_names": wb.sheetnames
        }

        logger.info(f"成功读取Excel文件: {file_path}, 工作表: {ws.title}")
        return True, result

    except Exception as e:
        logger.error(f"读取Excel文件失败: {e}")
        return False, {"error": f"读取Excel文件失败: {e}"}


def create_excel_file(file_path: Union[str, Path],
                      data: List[List[Any]] = None,
                      sheet_name: str = "Sheet1",
                      workspace_path: Optional[Union[str, Path]] = None,
                      overwrite: bool = False) -> Tuple[bool, str]:
    """创建新的Excel文件

    Args:
        file_path: 文件路径
        data: 初始数据（二维列表）
        sheet_name: 工作表名称
        workspace_path: 工作区路径
        overwrite: 是否覆盖已存在文件

    Returns:
        Tuple[bool, str]: 是否成功和消息
    """
    try:
        # 处理文件路径
        if isinstance(file_path, str):
            file_path = Path(file_path)

        # 处理工作区路径
        if workspace_path is not None:
            workspace = Path(workspace_path)
        else:
            workspace = Path.cwd()

        # 处理相对路径
        if not file_path.is_absolute():
            file_path = workspace / file_path

        # 确保文件在工作区内
        if not is_file_within_workspace(file_path, workspace):
            return False, "错误：文件路径不在当前工作区内"

        # 确保文件扩展名为.xlsx
        if file_path.suffix.lower() != ".xlsx":
            file_path = file_path.with_suffix(".xlsx")

        # 检查文件是否已存在
        if file_path.exists() and not overwrite:
            return False, f"错误：文件已存在: {file_path}。如需覆盖请设置 overwrite=True"

        # 创建父目录（如果不存在）
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 添加数据
        if data:
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        else:
            # 默认添加表头
            ws.cell(row=1, column=1, value="示例数据")

        # 保存文件
        wb.save(file_path)

        logger.info(f"成功创建Excel文件: {file_path}")
        return True, f"Excel文件创建成功: {file_path}"

    except Exception as e:
        return False, f"创建Excel文件失败: {e}"


def write_cell_to_excel(file_path: Union[str, Path],
                        cell_address: str,
                        value: Any,
                        sheet_name: Optional[str] = None,
                        workspace_path: Optional[Union[str, Path]] = None) -> Tuple[bool, str]:
    """向Excel单元格写入数据

    Args:
        file_path: Excel文件路径
        cell_address: 单元格地址（如 'A1', 'B2'）
        value: 要写入的值
        sheet_name: 工作表名称，None则使用活动工作表
        workspace_path: 工作区路径

    Returns:
        Tuple[bool, str]: 是否成功和消息
    """
    try:
        # 处理文件路径
        if isinstance(file_path, str):
            file_path = Path(file_path)

        # 处理工作区路径
        if workspace_path is not None:
            workspace = Path(workspace_path)
        else:
            workspace = Path.cwd()

        # 处理相对路径
        if not file_path.is_absolute():
            file_path = workspace / file_path

        # 安全检查
        if not is_file_within_workspace(file_path, workspace):
            return False, "错误：文件不在当前工作区内"

        if not file_path.exists():
            return False, f"错误：文件不存在: {file_path}"

        # 加载工作簿
        wb = load_workbook(filename=file_path)

        # 获取工作表
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                return False, f"错误：工作表 '{sheet_name}' 不存在"
        else:
            ws = wb.active

        # 写入单元格
        ws[cell_address] = value

        # 保存文件
        wb.save(file_path)

        logger.info(f"成功写入单元格 {cell_address}: {value}")
        return True, f"成功写入单元格 {cell_address}"

    except Exception as e:
        return False, f"写入单元格失败: {e}"


def read_cell_from_excel(file_path: Union[str, Path],
                         cell_address: str,
                         sheet_name: Optional[str] = None,
                         workspace_path: Optional[Union[str, Path]] = None) -> Tuple[bool, Any]:
    """读取Excel单元格数据

    Args:
        file_path: Excel文件路径
        cell_address: 单元格地址（如 'A1', 'B2'）
        sheet_name: 工作表名称，None则使用活动工作表
        workspace_path: 工作区路径

    Returns:
        Tuple[bool, Any]: 是否成功和单元格值
    """
    try:
        # 处理文件路径
        if isinstance(file_path, str):
            file_path = Path(file_path)

        # 处理工作区路径
        if workspace_path is not None:
            workspace = Path(workspace_path)
        else:
            workspace = Path.cwd()

        # 处理相对路径
        if not file_path.is_absolute():
            file_path = workspace / file_path

        # 安全检查
        if not is_file_within_workspace(file_path, workspace):
            return False, "错误：文件不在当前工作区内"

        if not file_path.exists():
            return False, f"错误：文件不存在: {file_path}"

        # 加载工作簿
        wb = load_workbook(filename=file_path, data_only=True)

        # 获取工作表
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                return False, f"错误：工作表 '{sheet_name}' 不存在"
        else:
            ws = wb.active

        # 读取单元格值
        cell_value = ws[cell_address].value

        logger.info(f"成功读取单元格 {cell_address}: {cell_value}")
        return True, cell_value

    except Exception as e:
        return False, f"读取单元格失败: {e}"


def add_sheet_to_excel(file_path: Union[str, Path],
                       sheet_name: str,
                       data: List[List[Any]] = None,
                       workspace_path: Optional[Union[str, Path]] = None) -> Tuple[bool, str]:
    """向Excel文件添加新工作表

    Args:
        file_path: Excel文件路径
        sheet_name: 新工作表名称
        data: 初始数据
        workspace_path: 工作区路径

    Returns:
        Tuple[bool, str]: 是否成功和消息
    """
    try:
        # 处理文件路径
        if isinstance(file_path, str):
            file_path = Path(file_path)

        # 处理工作区路径
        if workspace_path is not None:
            workspace = Path(workspace_path)
        else:
            workspace = Path.cwd()

        # 处理相对路径
        if not file_path.is_absolute():
            file_path = workspace / file_path

        # 安全检查
        if not is_file_within_workspace(file_path, workspace):
            return False, "错误：文件不在当前工作区内"

        if not file_path.exists():
            return False, f"错误：文件不存在: {file_path}"

        # 加载工作簿
        wb = load_workbook(filename=file_path)

        # 检查工作表名称是否已存在
        if sheet_name in wb.sheetnames:
            return False, f"错误：工作表 '{sheet_name}' 已存在"

        # 创建新工作表
        ws = wb.create_sheet(title=sheet_name)

        # 添加数据
        if data:
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # 保存文件
        wb.save(file_path)

        logger.info(f"成功添加工作表 '{sheet_name}' 到文件: {file_path}")
        return True, f"成功添加工作表 '{sheet_name}'"

    except Exception as e:
        return False, f"添加工作表失败: {e}"

if __name__ == '__main__':
    print('start..')
    status,data = read_excel_file('excel/发票信息汇总.xlsx','../../../workspace/workspace_admin')
    # status,data = create_excel_file('excel/新建表格.xlsx',[[1,2,3],[4,5,6]],'first_sheet','../../../workspace/workspace_admin')
    # status,data = write_cell_to_excel('excel/新建表格.xlsx','A1',100,'first_sheet','../../../workspace/workspace_admin')
    # status,data = read_cell_from_excel('excel/新建表格.xlsx','A1','first_sheet','../../../workspace/workspace_admin')
    # status,data = add_sheet_to_excel('excel/新建表格.xlsx','second_sheet',[[5,6,7],[8,9,10]],'../../../workspace/workspace_admin')
    print(data)
